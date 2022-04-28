from bs4 import BeautifulSoup
from urllib.request import urlopen as uReq
import requests
import speech_recognition as sr
import pyttsx3
import AppKit
import wikipedia
from sklearn.feature_extraction.text import TfidfVectorizer
import openpyxl
from openpyxl import load_workbook
import wikipedia
import wikipediaapi
import re, string
import heapq
import nltk
nltk.download('punkt')
nltk.download('stopwords')

def get_summary(url,q):
    title = url.split("/")[-1].split("_")
    #print(title)
    title = ' '.join(title)
    #print(title)
    wiki = wikipediaapi.Wikipedia('en')
    page = wiki.page(title)
    q = str.title(q)
    #print(q)
    secs = page.sections
    sec_dic = {}
    for s in page.sections:
        ss_list = []
        for ss in s.sections:
            ss_list.append(ss.title)
            #print(ss.title)
        sec_dic[s.title] = ss_list
        #print("###########")
    section = [s.title for s in page.sections]
    p = wikipedia.page(title) # Will take the Machine Learning URL
    sentences = p.section(q) 
    for ss in sec_dic[q]:
        sentences+= p.section(ss)
    print(sentences)
    print("#######################################################################################################################")
    article_text = re.sub(r'\[[0-9]*\]', ' ', sentences)
    article_text = re.sub(r'\s+', ' ', article_text)

    formatted_article_text = re.sub('[^a-zA-Z]', ' ', article_text )
    formatted_article_text = re.sub(r'\s+', ' ', formatted_article_text)


    sentence_list = nltk.sent_tokenize(article_text)
    stopwords = nltk.corpus.stopwords.words('english')
    #print(sentence_list)

    word_frequencies = {}
    for word in nltk.word_tokenize(formatted_article_text):
        if word not in stopwords:
            if word not in word_frequencies.keys():
                word_frequencies[word] = 1
            else:
                word_frequencies[word] += 1
        maximum_frequency = max(word_frequencies.values())

    for word in word_frequencies.keys():
        word_frequencies[word] = (word_frequencies[word]/maximum_frequency)
        sentence_scores = {}


    for sent in sentence_list:
        for word in nltk.word_tokenize(sent.lower()):
            if word in word_frequencies.keys():
                if len(sent.split(' ')) < 30:
                    if sent not in sentence_scores.keys():
                        sentence_scores[sent] = word_frequencies[word]
                    else:
                        sentence_scores[sent] += word_frequencies[word]


    summary_sentences = heapq.nlargest(7, sentence_scores, key=sentence_scores.get)

    summary = ' '.join(summary_sentences)
    print("#######################################################################################################################")
    print(summary)
    engine.say(summary)
    engine.runAndWait()


def get_intro(url):
    #url="https://en.wikipedia.org/wiki/"+link
        
    #print(url)
    page = requests.get(url)
    soup = BeautifulSoup(page.content, 'html.parser')
    intro=soup.select('p')
    for para in intro:
        i=para.get_text()
        if(len(i.split())) > 10:
            #print(i)
            return i

def a_tagModifier(soup,a_dic):
    fil = soup.find_all('a')
    for tag in fil:
        try:
            if tag.string != None :
                if tag.string[0] == '[':
                    tag.decompose()
                else:
                    a = len(tag.string)
                    a_dic[tag.string] = tag.get('href')
                    b = "$$"+tag.string+"$$"
                    tag.string = b
        except AttributeError:
            pass

def pparse(pagedic,url, a_dic):
    page = requests.get(url)
    soup = BeautifulSoup(page.content, 'html.parser')
    a_tagModifier(soup, a_dic)
    heading_tags = ["h1", "h2"]
    for section in soup.find_all(heading_tags):
        nextNode = section
        flag = True
        try:
            if nextNode.name == "h1":
                hed = nextNode.text
            else:
                hed = nextNode.span.get_text()
            bdy = ""
        except AttributeError:
            flag = False
        while flag:
            try:
                nextNode = nextNode.nextSibling
            except:
                break
            try:
                tag_name = nextNode.name
            except AttributeError:
                tag_name = ""
            try:
                bdy += nextNode.get_text()
            except AttributeError:
                pass
            if tag_name == "h2" or tag_name == "h1":
                pagedic[hed]=bdy
                break
            else:
                pass
            

def get_similarity(intro1, intro2):
    corpus = [intro1,intro2]
    vect = TfidfVectorizer(min_df=1, stop_words="english")                                                                                                                                                                                                   
    tfidf = vect.fit_transform(corpus)                                                                                                                                                                                                                       
    pairwise_similarity = tfidf * tfidf.T 
    pairwise_similarity = pairwise_similarity.toarray()   
    return pairwise_similarity[0][1]


def navigation(url):
    pagdic = {}
    a_dic = {}
    pparse(pagdic, url , a_dic)
    sheet_name = url.split("/")[-1]
    final_a_tags = []
    if sheet_name in wb.sheetnames:
        sheet = wb.get_sheet_by_name(sheet_name)
        first_column = sheet['A']
        for x in range(len(first_column)): 
            final_a_tags.append(first_column[x].value)
            #print() 
    else: 
        current_page_intro = get_intro(url)

        priority_a_tags = []
        for key in a_dic:
            #print(a_dic[key])
            next_page_intro = None
            if a_dic[key]!= None and a_dic[key].startswith("/wiki/"):
                next_page_intro = get_intro("https://en.wikipedia.org" + a_dic[key])
                if next_page_intro != None:
                    similarity = get_similarity(current_page_intro,next_page_intro)
                #print(next_page_intro)
                    ele = [key, similarity]
                    print(ele)
                    if ele not in priority_a_tags:
                        priority_a_tags.append(ele)

        priority_a_tags.sort(key = lambda x:x[1])
        required_size = len(priority_a_tags)//2
        print(required_size)
        del priority_a_tags[required_size:]
        #print(priority_a_tags)

        final_a_tags = [item[0] for item in priority_a_tags]
        sheet = wb.create_sheet(title=sheet_name)
        for item in final_a_tags:
            sheet.append([item])
        wb.save('a_tags_book.xlsx')

    curr = 0
    headings = list(pagdic.keys())
    #print(headings)

    while(curr!= len(headings)):
        heading = headings[curr]
        engine = pyttsx3.init() #tts engine
        engine.setProperty('voice', voices[0].id)
        engine.say(heading)
        engine.runAndWait()
        ques = ("Do you want to know about this?")
        engine.setProperty('voice', voices[0].id)
        engine.say(ques)
        engine.runAndWait()
        with sr.Microphone() as source:
            r.adjust_for_ambient_noise(source)
            AppKit.NSBeep()
            audio_data = r.record(source, duration=5)
            try:
            # using google speech recognition
                text = r.recognize_google(audio_data)
                if text == 'go back':
                    return 1
                if text == 'stop':
                    comm ="Thanks for using our application"
                    engine.setProperty('voice', voices[0].id)
                    engine.say(comm) 
                    engine.runAndWait()
                    return -1
                if text == 'no' or text == "next heading":
                    if(curr == len(headings) -1):
                        comm ="We have reached the end of the page."
                        engine.setProperty('voice', voices[0].id)
                        engine.say(comm) 
                        engine.runAndWait()
                    curr+=1
                if text == "previous heading":
                    if(curr == 0):
                        comm ="This is the first heading."
                        engine.setProperty('voice', voices[0].id)
                        engine.say(comm)
                        engine.runAndWait()   
                        continue          
                    curr-=1
                if text == "yes":
                    engine = pyttsx3.init() #tts engine
                    ques = ("Do you want to read the summary")
                    engine.setProperty('voice', voices[0].id)
                    engine.say(ques)
                    engine.runAndWait()
                    with sr.Microphone() as source:
                        r.adjust_for_ambient_noise(source)
                        AppKit.NSBeep()
                        audio_data = r.record(source, duration=5)
                        try:
                        # using google speech recognition
                            text = r.recognize_google(audio_data)
                            if text == 'yes':
                                get_summary(url,heading)
                        except:
                            comm=("Sorry, I did not get that")
                            engine.setProperty('voice', voices[0].id)
                            engine.say(comm)
                            engine.runAndWait()
                    engine = pyttsx3.init() #tts engine
                    ques = ("Do you want to read the full text")
                    engine.setProperty('voice', voices[0].id)
                    engine.say(ques)
                    engine.runAndWait()
                    with sr.Microphone() as source:
                        r.adjust_for_ambient_noise(source)
                        AppKit.NSBeep()
                        audio_data = r.record(source, duration=5)  
                        try:
                        # using google speech recognition
                            text = r.recognize_google(audio_data)
                            if text == 'yes':   
                                de_split = pagdic[heading].split("$$")
                                #print(de_split)
                                for po in range(0, len(de_split), 2):
                                    txt1 = de_split[po]
                                    txt2 = de_split[po+1]
                                    print(txt1, txt2)
                                    if(txt1 != ''):
                                        engine.setProperty('voice', voices[0].id)
                                        engine.say(txt1)
                                    engine.setProperty('voice', voices[0].id)
                                    engine.say(txt2)
                                    engine.runAndWait()
                                    if txt2 not in final_a_tags:
                                        continue
                                    ques = "Do you want to read more about " + txt2
                                    engine.setProperty('voice', voices[0].id)
                                    engine.say(ques)
                                    engine.runAndWait()
                                    with sr.Microphone() as source:
                                        r.adjust_for_ambient_noise(source)
                                        AppKit.NSBeep()
                                        audio_data = r.record(source, duration=5)
                                        try:
                                            # using google speech recognition
                                            text = r.recognize_google(audio_data)
                                            if text == 'no':
                                                continue
                                            elif text == 'go back':
                                                return 1
                                            elif text == 'stop':
                                                comm ="Thanks for using our application"
                                                engine.setProperty('voice', voices[0].id)
                                                engine.say(comm) 
                                                engine.runAndWait()
                                                return -1
                                            elif text == "next heading":
                                                if(curr == len(headings) -1):
                                                    comm ="We have reached the end of the page."
                                                    engine.setProperty('voice', voices[0].id)
                                                    engine.say(comm) 
                                                    engine.runAndWait()
                                                curr+=1
                                                break
                                            elif text == "previous heading":
                                                if(curr == 0):
                                                    comm ="This is the first heading."
                                                    engine.setProperty('voice', voices[0].id)
                                                    engine.say(comm)
                                                    engine.runAndWait()   
                                                    continue          
                                                curr-=1
                                                break
                                            elif text == 'yes':
                                                inp = "Do you want to read the summary " 
                                                engine.setProperty('voice', voices[0].id)
                                                engine.say(inp) 
                                                engine.runAndWait()
                                                while(True):
                                                    with sr.Microphone() as source:
                                                        r.adjust_for_ambient_noise(source)
                                                        AppKit.NSBeep()
                                                        audio_data = r.record(source, duration=5)
                                                        try:
                                                        # using google speech recognition
                                                            text = r.recognize_google(audio_data)
                                                            if text == 'no':
                                                                break
                                                            elif text == 'yes':
                                                                summ = wikipedia.summary(txt2)
                                                                engine.setProperty('voice', voices[0].id)
                                                                engine.say(summ) 
                                                                engine.runAndWait()
                                                                break
                                                        except:
                                                            comm=("Sorry, I did not get that. Please repeat")
                                                            engine.setProperty('voice', voices[0].id)
                                                            engine.say(comm)
                                                            engine.runAndWait()
                                                val = navigation("https://en.wikipedia.org" + a_dic[de_split[po+1]])
                                                if(val == -1):
                                                    return -1
                                                break
                                        except:
                                            comm=("Sorry, I did not get that")
                                            engine.setProperty('voice', voices[0].id)
                                            engine.say(comm)
                                            engine.runAndWait()
                                            po-=2
                        except:
                            comm=("Sorry, I did not get that")
                            engine.setProperty('voice', voices[0].id)
                            engine.say(comm)
                            engine.runAndWait()
            except:
                comm=("Sorry, I did not get that")
                engine.setProperty('voice', voices[0].id)
                engine.say(comm)
                engine.runAndWait()










import tkinter as tk
from tkinter import *
from urllib.request import urlopen
from PIL import Image, ImageTk
import io
import requests
import base64

r = tk.Tk()
r.geometry("400x400")

def strt(event=0):
    print("hfkd")
    label["text"] = "Listening..."
    b['state']= "disabled"
    bb['state']= tk.NORMAL
    wb = openpyxl.load_workbook('a_tags_book.xlsx')  #/Users/devanshigarg/Desktop/Major Project/
#print("Workbook", wb.sheetnames) 

    r = sr.Recognizer()
    S = requests.Session()

    engine = pyttsx3.init() #tts engine
    voices = engine.getProperty('voices')
    inp = "Say the topic to be searched " 
    engine.setProperty('voice', voices[0].id)
    engine.say(inp) 
    engine.runAndWait()
    with sr.Microphone() as source:
        r.adjust_for_ambient_noise(source)
        AppKit.NSBeep()
        audio_data = r.record(source, duration=5)
        try:
            # using google speech recognition
            search = r.recognize_google(audio_data)
        except:
            comm=("Sorry, I did not get that")
            engine.setProperty('voice', voices[0].id)
            engine.say(comm)
            engine.runAndWait()
    topic = search.split(" ")

    url = "https://en.wikipedia.org/wiki/"

    for i in range(0, len(topic)):
        if i==0:
            url+= topic[i]
        else:
            url += "_" + topic[i]

    inp = "Do you want to read the summary " 
    engine.setProperty('voice', voices[0].id)
    engine.say(inp) 
    engine.runAndWait()

    while(True):
        with sr.Microphone() as source:
            r.adjust_for_ambient_noise(source)
            AppKit.NSBeep()
            audio_data = r.record(source, duration=5)
            try:
            # using google speech recognition
                text = r.recognize_google(audio_data)
                if text == 'no':
                    break
                if text == 'yes':
                    summ = wikipedia.summary(search)
                    engine.setProperty('voice', voices[0].id)
                    engine.say(summ) 
                    engine.runAndWait()
                    break
            except:
                comm=("Sorry, I did not get that. Please repeat")
                engine.setProperty('voice', voices[0].id)
                engine.say(comm)
                engine.runAndWait()


    navigation(url)


def stpp(event=0):
    label["text"] = "Please Press Start"
    b['state']= tk.NORMAL
    bb["state"]= "disabled"
    


Label(r, text=" Listening : "
     ).grid(row=0, sticky=W, rowspan=5)
  
  
b = Button(r, text="Start", command=strt)
b.grid(row=3, column=2, columnspan=2, rowspan=2,
       padx=5, pady=5)


bb = Button(r, text="Stop", command=stpp)
bb.grid(row=3, column=8, columnspan=2, rowspan=2,
       padx=5, pady=5)
bb['state']="disabled"

label =Label(r, text=" Please Press Start ")
label.grid(row=6,column=2)
Label(r, text=" To start press a ")
label.grid(row=7,column=2)
Label(r, text=" To stop press z ")
label.grid(row=8,column=2)

r.title('Wikipedia For Blind')

r.bind('<a>', strt)
r.bind('<z>', stpp)



r.mainloop()